# 크롬 브라우저를 띄우기 위해, 웹드라이버를 가져오기
import chromedriver_autoinstaller
chromedriver_autoinstaller.install()

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from openpyxl import Workbook, load_workbook #excel
import datetime
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'worksheet'
ws2 = wb.create_sheet('worksheet2')
ws2['B2'] = 'cell'




# 크롬 드라이버로 크롬을 실행한다.
try:
    # 입찰정보 검색 페이지로 이동
    driver = webdriver.Chrome()
    driver.get('https://www.g2b.go.kr:8101/ep/tbid/tbidFwd.do')

#tbid = 입찰공고
#taskClCds = 업무구분
#bidNm = 공고명
#fromBidDt : 공고일 ~부터 2015%2F09%2F01
#toBidDt : 공고일 ~까지
#fromOpenBidDt=&toOpenBidDt= 개찰일 ~에서 ~까지
#exceptEnd=Y 입찰마감건제외
#instNm : 공공기관
#radOrgan : 1(공고기관) 2(수요기관)
#area : 지역[00 전국(지역제한없음, 30 대전 )]
#industryCd= : 업종 [1162, 1164,.. . . .]
#bidno : 공고번호
#recordCountPerPage = 100 : 목록수 100개


    # 업무 종류 체크
    task_dict = {'용역': 'taskClCds5'}
#    task_dict = {'용역': 'taskClCds5', '민간': 'taskClCds20', '기타': 'taskClCds4'}

    for task in task_dict.values():
        checkbox = driver.find_element_by_id(task)
        checkbox.click()

    # 검색어
    query = "경비"
    # id값이 bidNm인 태그 가져오기
    bidNm = driver.find_element_by_id('bidNm')
    # 내용을 삭제 (버릇처럼 사용할 것!)
    bidNm.clear()
    # 검색어 입력후 엔터
    bidNm.send_keys(query)
    bidNm.send_keys(Keys.RETURN)

    # 검색 조건 체크
    option_dict = {'검색기간 1달': 'setMonth1_1', '입찰마감건 제외': 'exceptEnd', '검색건수 표시': 'useTotalCount'}
    for option in option_dict.values():
        checkbox = driver.find_element_by_id(option)
        checkbox.click()

    # 목록수 100건 선택 (드롭다운)
    recordcountperpage = driver.find_element_by_name('recordCountPerPage')
    selector = Select(recordcountperpage)
    selector.select_by_value('100')

    # 검색 버튼 클릭
    search_button = driver.find_element_by_class_name('btn_mdl')
    search_button.click()

    # 검색 결과 확인
    elem = driver.find_element_by_class_name('results')
    div_list = elem.find_elements_by_tag_name('div')

    # 검색 결과 모두 긁어서 리스트로 저장
    results = []
    for div in div_list:
        results.append(div.text)
        a_tags = div.find_elements_by_tag_name('a')
        if a_tags:
            for a_tag in a_tags:
                link = a_tag.get_attribute('href')
                results.append(link)

    # 검색결과 모음 리스트를 12개씩 분할하여 새로운 리스트로 저장
    #result = [results[i * 12:(i + 1) * 12] for i in range((len(results) + 12 - 1) // 12 )]
#now = datetime.datetime.now()
#date = now.strftime('%Y.%m.%d')
#excel_file_path = 'C:\Users\USER\AppData\Local\Packages\PythonSoftwareFoundation.Python.3.9_qbz5n2kfra8p0\LocalCache\local-packages\Python39\Scripts\htdocs\NARA'
#excel_file_path = 'c:\test'

#excel 파일 저장위치
#excel_file_name = excel_file_path + date + '.xlsx' #엑셀파일이름
#filename ="test.xlsx"
#wb.save(filename)


#filename = 'test.xlsx'
#excel_sheet_title = 'comfirm'
#excel_row = 2

def make_excel():
    work_book = Workbook()
    sheet1 = work_book.active
    sheet1.title = excel_sheet_title

    #헤더입력
    sheet1.cell(row=1, column=1).value = '업무'
    sheet1.cell(row=1, column=2).value = '공고번호'
    sheet1.cell(row=1, column=3).value = '투찰'
    sheet1.cell(row=1, column=4).value = '분류'
    sheet1.cell(row=1, column=5).value = '공고명'
    sheet1.cell(row=1, column=6).value = '공고기관'
    sheet1.cell(row=1, column=7).value = '수요기관'
    sheet1.cell(row=1, column=8).value = '계약방법'
    sheet1.cell(row=1, column=9).value = '입력일시및마감'
    sheet1.cell(row=1, column=10).value = '공동수급'
    sheet1.cell(row=1, column=11).value = '비고'
    sheet1.cell(row=1, column=12).value = '업무'

    work_book.save(filename=excel_file_name)
    work_book.close()

def insert_data_to_excel(result):
    excel_file = load_workbook(excel_file_name)
    sheet1 = excel_file[excel_sheet_title]

    excel_column = 3
    for data in result:
        sheet1.cell(row=excel_row, column=excel_column).value = data
        excel_column +=1

    excel_file.save(excel_file_name)
    excel_file.close()

    # 결과 출력
#    print(result)

except Exception as e:
    # 위 코드에서 에러가 발생한 경우 출력
    print(e)

finally:
#    input("Press enter to exit ;)")
    # 에러와 관계없이 실행되고, 크롬 드라이버를 종료
    driver.quit()
